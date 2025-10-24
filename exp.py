import time
import pandas as pd
from utils import lp_cvxp
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from datetime import timedelta


# COPT HiGHS CBC IPOPT
# === ПАРАМЕТРЫ ЭКСПЕРИМЕНТА ===
sizes = [500, 1000, 1500, 2000]
density = 0.01
solvers = ["CLARABEL", "HIGHS", "OSQP", "ECOS", "ECOS_BB"]
output_xlsx = "lp_solver_times_pivot.xlsx"

rows = []

# === ОСНОВНОЙ ЦИКЛ ЭКСПЕРИМЕНТА ===
for n in sizes:
    print(f"\n=== Размер задачи: n = m = {n} ===")
    prob, A, b, c, x = lp_cvxp(n=n, m=n, density=density, easy=False, seed=42)

    for solver in solvers:
        print(f"  → Решаем с {solver:<9} ...", end=" ", flush=True)
        t0 = time.monotonic()
        try:
            prob.solve(solver=solver, verbose=False)
            elapsed = time.monotonic() - t0
            status = prob.status
            obj_val = prob.value
            print(f"{elapsed:7.3f} с, статус={status}")
        except Exception as e:
            elapsed = None
            status = "error"
            obj_val = None
            print(f"ошибка: {type(e).__name__}")

        rows.append(
            {
                "n": n,
                "solver": solver,
                "status": status,
                "time_sec": round(elapsed, 6) if elapsed else None,
                "objective": obj_val,
            }
        )

# === СОЗДАЁМ DATAFRAME ===
df = pd.DataFrame(rows)


# --- Форматируем секунды в Ч:ММ:СС.мс ---
def fmt_time(sec):
    if sec is None or pd.isna(sec):
        return "—"
    td = timedelta(seconds=float(sec))
    total_seconds = td.total_seconds()
    hours, remainder = divmod(int(total_seconds), 3600)
    minutes, seconds = divmod(int(remainder), 60)
    milliseconds = int((total_seconds - int(total_seconds)) * 1000)
    return f"{hours}:{minutes:02d}:{seconds:02d}.{milliseconds:03d}"


df["time_str"] = df["time_sec"].apply(fmt_time)

# === СВОДНЫЕ ТАБЛИЦЫ ===
pivot_time = df.pivot(index="n", columns="solver", values="time_str").reset_index()
pivot_status = df.pivot(index="n", columns="solver", values="status").reset_index()

# === СОХРАНЯЕМ С ПОДСВЕТКОЙ ===
wb = Workbook()
ws = wb.active
ws.title = "time_sec"

for r in dataframe_to_rows(pivot_time, index=False, header=True):
    ws.append(r)

# --- Цветовая шкала по времени ---
for row in ws.iter_rows(min_row=2, min_col=2):
    for cell in row:
        val = cell.value
        if not val or val == "—":
            continue
        try:
            time_part = val.split(".")[0]  # отбрасываем миллисекунды
            h, m, s = map(int, time_part.split(":"))
            total_sec = h * 3600 + m * 60 + s
        except Exception:
            total_sec = 0

        # Простая дискретная шкала
        if total_sec < 2:
            fill = PatternFill("solid", fgColor="C6EFCE")  # зелёный
        elif total_sec < 10:
            fill = PatternFill("solid", fgColor="FFEB9C")  # жёлтый
        elif total_sec < 60:
            fill = PatternFill("solid", fgColor="FFC7CE")  # розовый
        else:
            fill = PatternFill("solid", fgColor="FF6666")  # красный
        cell.fill = fill

# === ДОБАВЛЯЕМ ЛИСТ СО СТАТУСАМИ ===
ws2 = wb.create_sheet("status")
for r in dataframe_to_rows(pivot_status, index=False, header=True):
    ws2.append(r)

wb.save(output_xlsx)
print(f"\n✅ Результаты сохранены в файл: {output_xlsx}")
